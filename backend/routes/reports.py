from flask import Blueprint, request, jsonify, make_response
from flask_login import login_required, current_user
from auth_utils import token_required, account_filter, expenses_for_tenant
from models import db, Sale, SaleItem, Product, Category, Customer, Expense, User
from sqlalchemy import func
from datetime import datetime, date, timedelta
import csv, io

reports_bp = Blueprint("reports", __name__)

def get_date_range():
    date_from = request.args.get("date_from", date.today().replace(day=1).isoformat())
    date_to   = request.args.get("date_to",   date.today().isoformat())
    return date_from, date_to + "T23:59:59"


def _sales_q():
    return account_filter(Sale.query, Sale)


def _products_q():
    return account_filter(Product.query, Product)


def _users_q():
    return account_filter(User.query, User)


def _completed_sales_in_range(date_from, date_to):
    d_from = date_from[:10]
    d_to = date_to[:10]
    return _sales_q().filter(
        func.date(Sale.sale_date) >= d_from,
        func.date(Sale.sale_date) <= d_to,
        Sale.status == "completed",
    )


def _expenses_in_range(date_from, date_to):
    return expenses_for_tenant().filter(
        Expense.expense_date >= date_from[:10],
        Expense.expense_date <= date_to[:10],
    )


# ── Daily ─────────────────────────────────────────────────────────────────────

@reports_bp.route("/daily", methods=["GET"])
@token_required
@login_required
def daily_report():
    target = request.args.get("date", date.today().isoformat())
    sales = _sales_q().filter(
        func.date(Sale.sale_date) == target, Sale.status == "completed"
    ).all()
    total_revenue = sum(float(s.total) for s in sales)
    total_profit  = sum(
        sum((float(i.unit_price) - float(i.cost_price)) * i.qty for i in s.items)
        for s in sales
    )
    return jsonify({
        "date":         target,
        "transactions": len(sales),
        "revenue":      total_revenue,
        "profit":       total_profit,
        "sales":        [s.to_dict(include_items=True) for s in sales],
    })


# ── P&L Summary ───────────────────────────────────────────────────────────────

@reports_bp.route("/summary", methods=["GET"])
@token_required
@login_required
def summary_report():
    date_from, date_to = get_date_range()
    sales = _completed_sales_in_range(date_from, date_to).all()
    expenses = _expenses_in_range(date_from, date_to).all()
    total_revenue  = sum(float(s.total) for s in sales)
    total_cost     = sum(float(i.cost_price) * i.qty for s in sales for i in s.items)
    gross_profit   = total_revenue - total_cost
    total_expenses = sum(float(e.amount) for e in expenses)
    net_profit     = gross_profit - total_expenses

    by_payment = {}
    for s in sales:
        pm = s.payment_method
        by_payment[pm] = by_payment.get(pm, 0) + float(s.total)

    daily = {}
    for s in sales:
        day = s.sale_date.date().isoformat() if s.sale_date else None
        if day:
            daily[day] = daily.get(day, 0) + float(s.total)
    trend = [{"date": d, "revenue": v} for d, v in sorted(daily.items())]

    return jsonify({
        "date_from":         date_from[:10],
        "date_to":           date_to[:10],
        "transactions":      len(sales),
        "total_revenue":     total_revenue,
        "total_cost":        total_cost,
        "gross_profit":      gross_profit,
        "total_expenses":    total_expenses,
        "net_profit":        net_profit,
        "gross_margin_pct":  round((gross_profit / total_revenue * 100) if total_revenue > 0 else 0, 2),
        "payment_breakdown": by_payment,
        "daily_trend":       trend,
    })


# ── Inventory ─────────────────────────────────────────────────────────────────

@reports_bp.route("/inventory", methods=["GET"])
@token_required
@login_required
def inventory_report():
    products = _products_q().filter_by(status="active").all()
    items = [{
        "id":            p.id,
        "name":          p.name,
        "sku":           p.sku,
        "category":      p.category.name if p.category else "—",
        "cost_price":    float(p.cost_price or 0),
        "selling_price": float(p.selling_price or 0),
        "stock_qty":     p.stock_qty,
        "reorder_level": p.reorder_level,
        "is_low_stock":  p.is_low_stock,
        "cost_value":    float(p.cost_price or 0) * p.stock_qty,
        "retail_value":  float(p.selling_price or 0) * p.stock_qty,
    } for p in products]
    total_cost_value   = sum(i["cost_value"]   for i in items)
    total_retail_value = sum(i["retail_value"] for i in items)
    return jsonify({
        "items":              items,
        "total_cost_value":   total_cost_value,
        "total_retail_value": total_retail_value,
        "potential_profit":   total_retail_value - total_cost_value,
        "low_stock_count":    sum(1 for i in items if i["is_low_stock"]),
    })


# ── Product Performance ───────────────────────────────────────────────────────

@reports_bp.route("/products", methods=["GET"])
@token_required
@login_required
def product_performance():
    date_from, date_to = get_date_range()
    in_range = _completed_sales_in_range(date_from, date_to)
    rows = (
        db.session.query(
            SaleItem.product_id,
            func.sum(SaleItem.qty).label("units_sold"),
            func.sum(SaleItem.qty * SaleItem.unit_price).label("revenue"),
            func.sum(SaleItem.qty * SaleItem.cost_price).label("cost"),
        )
        .join(Sale, Sale.id == SaleItem.sale_id)
        .filter(Sale.id.in_(in_range.with_entities(Sale.id).scalar_subquery()))
        .group_by(SaleItem.product_id)
        .all()
    )

    products = {p.id: p for p in _products_q().all()}
    result = []
    for r in rows:
        p = products.get(r.product_id)
        revenue = float(r.revenue or 0)
        cost    = float(r.cost or 0)
        profit  = revenue - cost
        result.append({
            "product_id":   r.product_id,
            "name":         p.name if p else f"#{r.product_id}",
            "sku":          p.sku if p else "",
            "category":     p.category.name if p and p.category else "—",
            "units_sold":   int(r.units_sold or 0),
            "revenue":      round(revenue, 2),
            "cost":         round(cost, 2),
            "gross_profit": round(profit, 2),
            "margin_pct":   round((profit / revenue * 100) if revenue > 0 else 0, 2),
            "stock_qty":    p.stock_qty if p else 0,
        })

    result.sort(key=lambda x: x["revenue"], reverse=True)
    return jsonify({
        "date_from": date_from[:10],
        "date_to":   date_to[:10],
        "products":  result,
    })


# ── Staff Performance ─────────────────────────────────────────────────────────

@reports_bp.route("/staff", methods=["GET"])
@token_required
@login_required
def staff_performance():
    if current_user.role not in ["owner", "superadmin", "manager"]:
        return jsonify({"error": "Forbidden"}), 403

    date_from, date_to = get_date_range()
    in_range = _completed_sales_in_range(date_from, date_to)
    rows = (
        db.session.query(
            Sale.cashier_id,
            func.count(Sale.id).label("transactions"),
            func.sum(Sale.total).label("revenue"),
            func.avg(Sale.total).label("avg_sale"),
        )
        .filter(Sale.id.in_(in_range.with_entities(Sale.id).scalar_subquery()))
        .group_by(Sale.cashier_id)
        .all()
    )

    users = {u.id: u for u in _users_q().all()}
    result = []
    for r in rows:
        u = users.get(r.cashier_id)
        result.append({
            "cashier_id":   r.cashier_id,
            "cashier_name": u.full_name if u else f"User #{r.cashier_id}",
            "role":         u.role if u else "—",
            "transactions": int(r.transactions or 0),
            "revenue":      round(float(r.revenue or 0), 2),
            "avg_sale":     round(float(r.avg_sale or 0), 2),
        })

    result.sort(key=lambda x: x["revenue"], reverse=True)
    return jsonify({
        "date_from": date_from[:10],
        "date_to":   date_to[:10],
        "staff":     result,
    })


# ── CSV Export ────────────────────────────────────────────────────────────────

@reports_bp.route("/export/csv", methods=["GET"])
@token_required
@login_required
def export_csv():
    report_type = request.args.get("type", "sales")
    date_from, date_to = get_date_range()
    output = io.StringIO()
    writer = csv.writer(output)

    if report_type == "sales":
        writer.writerow(["Invoice", "Date", "Customer", "Cashier", "Subtotal",
                         "Discount", "Tax", "Total", "Payment"])
        sales = _completed_sales_in_range(date_from, date_to).all()
        for s in sales:
            writer.writerow([
                s.invoice_number, s.sale_date,
                s.customer.name if s.customer else "Walk-in",
                s.cashier.full_name if s.cashier else "",
                s.subtotal, s.discount_amount, s.tax_amount, s.total, s.payment_method,
            ])
    elif report_type == "inventory":
        writer.writerow(["SKU", "Name", "Category", "Cost", "Price", "Stock", "Value"])
        for p in _products_q().filter_by(status="active").all():
            writer.writerow([p.sku, p.name, p.category.name if p.category else "",
                             p.cost_price, p.selling_price, p.stock_qty,
                             float(p.selling_price or 0) * p.stock_qty])

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv"
    response.headers["Content-Disposition"] = f"attachment; filename={report_type}_report.csv"
    return response


# ── Excel Export (multi-sheet) ────────────────────────────────────────────────

@reports_bp.route("/export/xlsx", methods=["GET"])
@token_required
@login_required
def export_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed — run: pip install openpyxl"}), 500

    date_from, date_to = get_date_range()

    sales = _completed_sales_in_range(date_from, date_to).order_by(Sale.sale_date).all()
    expenses = _expenses_in_range(date_from, date_to).all()

    wb = openpyxl.Workbook()

    HEADER = "FF1A2332"
    WHITE  = "FFFFFFFF"
    GOLD   = "FFD4AF37"
    LIGHT  = "FFF0F4F8"
    YELLOW = "FFFDE68A"

    hdr_font = Font(bold=True, color=WHITE,  name="Calibri", size=10)
    hdr_fill = PatternFill("solid", fgColor=HEADER)

    def style_headers(ws, row, n):
        for col in range(1, n + 1):
            ws.cell(row=row, column=col).font = hdr_font
            ws.cell(row=row, column=col).fill = hdr_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    def auto_width(ws):
        for col in ws.columns:
            ltr = get_column_letter(col[0].column)
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[ltr].width = min(w + 4, 40)

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Your Store — Business Report"
    ws1["A1"].font = Font(bold=True, color=GOLD, name="Calibri", size=14)
    ws1["A2"] = f"Period: {date_from[:10]}  to  {date_to[:10]}"
    ws1["A2"].font = Font(italic=True, color="FF888888", size=9)

    total_revenue  = sum(float(s.total) for s in sales)
    total_cost     = sum(float(i.cost_price) * i.qty for s in sales for i in s.items)
    gross_profit   = total_revenue - total_cost
    total_expenses = sum(float(e.amount) for e in expenses)
    net_profit     = gross_profit - total_expenses

    kpis = [
        ("Metric", "Value"),
        ("Total Revenue",  total_revenue),
        ("Cost of Goods",  total_cost),
        ("Gross Profit",   gross_profit),
        ("Gross Margin %", round((gross_profit / total_revenue * 100) if total_revenue > 0 else 0, 1)),
        ("Total Expenses", total_expenses),
        ("Net Profit",     net_profit),
        ("Transactions",   len(sales)),
    ]
    for i, (label, val) in enumerate(kpis, start=4):
        ws1.cell(row=i, column=1, value=label)
        c = ws1.cell(row=i, column=2, value=val)
        if isinstance(val, float):
            c.number_format = '#,##0.00'
        if label == "Metric":
            style_headers(ws1, i, 2)

    by_payment = {}
    for s in sales:
        by_payment[s.payment_method] = by_payment.get(s.payment_method, 0) + float(s.total)

    row_pm = len(kpis) + 7
    ws1.cell(row=row_pm, column=1, value="Payment Breakdown").font = Font(bold=True, name="Calibri")
    ws1.cell(row=row_pm + 1, column=1, value="Method")
    ws1.cell(row=row_pm + 1, column=2, value="Amount")
    style_headers(ws1, row_pm + 1, 2)
    for j, (method, amount) in enumerate(by_payment.items(), start=row_pm + 2):
        ws1.cell(row=j, column=1, value=method.capitalize())
        ws1.cell(row=j, column=2, value=amount).number_format = '#,##0.00'

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Sales ──
    ws2 = wb.create_sheet("Sales")
    h2 = ["Invoice", "Date", "Customer", "Cashier", "Payment",
          "Subtotal", "Discount", "Tax", "Total", "Status"]
    for col, h in enumerate(h2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_headers(ws2, 1, len(h2))
    for ri, s in enumerate(sales, start=2):
        row_vals = [
            s.invoice_number,
            s.sale_date.strftime("%Y-%m-%d %H:%M") if s.sale_date else "",
            s.customer.name if s.customer else "Walk-in",
            s.cashier.full_name if s.cashier else "",
            s.payment_method,
            float(s.subtotal or 0), float(s.discount_amount or 0),
            float(s.tax_amount or 0), float(s.total or 0),
            s.status,
        ]
        for ci, val in enumerate(row_vals, start=1):
            c = ws2.cell(row=ri, column=ci, value=val)
            if ci in (6, 7, 8, 9):
                c.number_format = '#,##0.00'
        if ri % 2 == 0:
            for ci in range(1, 11):
                ws2.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=LIGHT)
    auto_width(ws2)

    # ── Sheet 3: Product Performance ──
    ws3 = wb.create_sheet("Product Performance")
    h3 = ["Product", "SKU", "Category", "Units Sold", "Revenue", "Cost", "Gross Profit", "Margin %"]
    for col, h in enumerate(h3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_headers(ws3, 1, len(h3))

    in_range = _completed_sales_in_range(date_from, date_to)
    prod_rows = (
        db.session.query(
            SaleItem.product_id,
            func.sum(SaleItem.qty).label("units"),
            func.sum(SaleItem.qty * SaleItem.unit_price).label("revenue"),
            func.sum(SaleItem.qty * SaleItem.cost_price).label("cost"),
        )
        .join(Sale, Sale.id == SaleItem.sale_id)
        .filter(Sale.id.in_(in_range.with_entities(Sale.id).scalar_subquery()))
        .group_by(SaleItem.product_id)
        .all()
    )
    products_map = {p.id: p for p in _products_q().all()}
    prod_data = sorted([
        (
            (products_map[r.product_id].name if r.product_id in products_map else f"#{r.product_id}"),
            (products_map[r.product_id].sku if r.product_id in products_map else ""),
            (products_map[r.product_id].category.name if r.product_id in products_map and products_map[r.product_id].category else "—"),
            int(r.units or 0),
            float(r.revenue or 0),
            float(r.cost or 0),
            float(r.revenue or 0) - float(r.cost or 0),
            round(((float(r.revenue or 0) - float(r.cost or 0)) / float(r.revenue) * 100) if r.revenue else 0, 1),
        )
        for r in prod_rows
    ], key=lambda x: x[4], reverse=True)

    for ri, row in enumerate(prod_data, start=2):
        for ci, val in enumerate(row, start=1):
            c = ws3.cell(row=ri, column=ci, value=val)
            if ci in (5, 6, 7):
                c.number_format = '#,##0.00'
        if ri % 2 == 0:
            for ci in range(1, 9):
                ws3.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=LIGHT)
    auto_width(ws3)

    # ── Sheet 4: Staff Performance ──
    ws4 = wb.create_sheet("Staff Performance")
    h4 = ["Cashier", "Role", "Transactions", "Total Revenue", "Avg Sale Value"]
    for col, h in enumerate(h4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_headers(ws4, 1, len(h4))

    staff_rows = (
        db.session.query(
            Sale.cashier_id,
            func.count(Sale.id).label("txns"),
            func.sum(Sale.total).label("revenue"),
            func.avg(Sale.total).label("avg_sale"),
        )
        .filter(Sale.id.in_(in_range.with_entities(Sale.id).scalar_subquery()))
        .group_by(Sale.cashier_id)
        .all()
    )
    users_map = {u.id: u for u in _users_q().all()}
    staff_data = sorted([
        (
            users_map[r.cashier_id].full_name if r.cashier_id in users_map else f"User #{r.cashier_id}",
            users_map[r.cashier_id].role if r.cashier_id in users_map else "—",
            int(r.txns or 0),
            float(r.revenue or 0),
            float(r.avg_sale or 0),
        )
        for r in staff_rows
    ], key=lambda x: x[3], reverse=True)

    for ri, row in enumerate(staff_data, start=2):
        for ci, val in enumerate(row, start=1):
            c = ws4.cell(row=ri, column=ci, value=val)
            if ci in (4, 5):
                c.number_format = '#,##0.00'
        if ri % 2 == 0:
            for ci in range(1, 6):
                ws4.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=LIGHT)
    auto_width(ws4)

    # ── Sheet 5: Inventory ──
    ws5 = wb.create_sheet("Inventory")
    h5 = ["SKU", "Name", "Category", "Cost Price", "Selling Price",
          "Stock Qty", "Reorder Level", "Cost Value", "Retail Value", "Status"]
    for col, h in enumerate(h5, 1):
        ws5.cell(row=1, column=col, value=h)
    style_headers(ws5, 1, len(h5))

    for ri, p in enumerate(_products_q().order_by(Product.name).all(), start=2):
        vals = [
            p.sku or "", p.name,
            p.category.name if p.category else "—",
            float(p.cost_price or 0), float(p.selling_price or 0),
            p.stock_qty, p.reorder_level,
            float(p.cost_price or 0) * p.stock_qty,
            float(p.selling_price or 0) * p.stock_qty,
            p.status,
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws5.cell(row=ri, column=ci, value=val)
            if ci in (4, 5, 8, 9):
                c.number_format = '#,##0.00'
        if p.is_low_stock:
            ws5.cell(row=ri, column=6).fill = PatternFill("solid", fgColor=YELLOW)
        elif ri % 2 == 0:
            for ci in range(1, 11):
                ws5.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=LIGHT)
    auto_width(ws5)

    # ── Serialize ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"retailos_report_{date_from[:10]}_to_{date_to[:10]}.xlsx"
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp
