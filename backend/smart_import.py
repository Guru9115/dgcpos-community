"""
Smart Product Import Engine — CSV, Excel (.xlsx/.xls), and SVC (delimited) files.
Auto-maps columns and creates categories from store-type aware imports.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from datetime import datetime

from models import db, Product, Category, InventoryMovement, Setting
from store_engine import normalize_store_type, get_category_names, get_store_config
from sqlalchemy import insert

MAX_BYTES = 10 * 1024 * 1024  # 10 MB

COLUMN_ALIASES: dict[str, list[str]] = {
    "name": [
        "name", "product", "product name", "product_name", "item", "item name",
        "item_name", "title", "description", "product description", "goods",
    ],
    "sku": ["sku", "code", "item code", "item_code", "product code", "product_code", "barcode sku"],
    "barcode": ["barcode", "ean", "upc", "isbn", "scan code", "scan_code"],
    "category": ["category", "cat", "type", "department", "group", "section", "class"],
    "cost_price": ["cost", "cost price", "cost_price", "purchase price", "buy price", "cp"],
    "selling_price": ["price", "selling price", "selling_price", "sale price", "mrp", "rsp", "sp"],
    "stock_qty": ["stock", "stock qty", "stock_qty", "quantity", "qty", "on hand", "on_hand"],
    "reorder_level": ["reorder", "reorder level", "reorder_level", "min stock", "min_stock"],
    "unit": ["unit", "uom", "measure"],
    "description": ["notes", "detail", "remarks", "long description"],
    "status": ["status", "active"],
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _map_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Map arbitrary column headers to canonical product fields."""
    normalized = {_norm_header(k): (v if v is not None else "") for k, v in raw.items()}
    out: dict[str, Any] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                val = normalized[alias]
                if val != "":
                    out[field] = val
                break
    return out


def _detect_delimiter(sample: str) -> str:
    for delim in (";", "\t", "|", ","):
        if sample.count(delim) >= 2:
            return delim
    return ","


def parse_csv_text(text: str) -> list[dict[str, Any]]:
    sample = text[:2048]
    delim = _detect_delimiter(sample)
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    return [dict(row) for row in reader]


def parse_excel_bytes(data: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(row):
            continue
        raw = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        out.append(raw)
    return out


def parse_file(filename: str, raw_bytes: bytes) -> list[dict[str, Any]]:
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        return parse_excel_bytes(raw_bytes)
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    return parse_csv_text(text)


def _get_or_create_category(name: str, account_id: int | None) -> Category:
    name = name.strip()[:64]
    if not name:
        raise ValueError("empty category")
    cat = Category.query.filter_by(name=name).first()
    if not cat:
        cat = Category(name=name)
        db.session.add(cat)
        db.session.flush()
    return cat


def _guess_category(name: str, store_type: str, known: set[str]) -> str | None:
    """Suggest category from product name keywords for store type."""
    n = name.lower()
    cfg = get_store_config(store_type)
    for cat in cfg.get("categories", []):
        if cat.lower() in n:
            return cat
    for cat in known:
        if cat.lower() in n:
            return cat
    return None


def import_product_rows(
    account_id: int,
    rows: list[dict[str, Any]],
    *,
    store_type: str | None = None,
    default_unit: str | None = None,
    created_by: int | None = None,
) -> dict[str, Any]:
    store_type = normalize_store_type(store_type)
    cfg = get_store_config(store_type)
    unit_default = default_unit or cfg.get("default_unit", "pcs")

    created = updated = skipped = 0
    categories_created: set[str] = set()
    errors: list[dict[str, Any]] = []

    # Pre-seed store categories if account has none
    existing_cats = {c.name for c in Category.query.all()}
    for cat_name in get_category_names(store_type):
        if cat_name not in existing_cats:
            db.session.add(Category(name=cat_name))
            categories_created.add(cat_name)
            existing_cats.add(cat_name)
    if categories_created:
        db.session.flush()

    for i, raw in enumerate(rows, start=2):
        row = _map_row(raw)
        name = str(row.get("name") or "").strip()
        if not name:
            skipped += 1
            errors.append({"row": i, "error": "Missing product name"})
            continue
        try:
            sku = str(row.get("sku") or "").strip() or None
            barcode = str(row.get("barcode") or "").strip() or None
            cat_name = str(row.get("category") or "").strip()
            if not cat_name:
                cat_name = _guess_category(name, store_type, existing_cats) or ""
            cost_price = float(row.get("cost_price") or 0)
            selling_price = float(row.get("selling_price") or row.get("price") or 0)
            stock_qty = int(float(row.get("stock_qty") or 0))
            reorder_level = int(float(row.get("reorder_level") or 5))
            unit = str(row.get("unit") or unit_default).strip() or unit_default
            status = str(row.get("status") or "active").strip().lower()
            if status not in ("active", "inactive"):
                status = "active"
            description = str(row.get("description") or "").strip() or None

            cat_id = None
            if cat_name:
                cat = _get_or_create_category(cat_name, account_id)
                cat_id = cat.id
                if cat_name not in existing_cats:
                    categories_created.add(cat_name)
                    existing_cats.add(cat_name)

            existing = None
            if sku:
                existing = Product.query.filter_by(sku=sku, account_id=account_id).first()
            if not existing:
                existing = Product.query.filter_by(name=name, account_id=account_id).first()

            if existing:
                existing.name = name
                existing.cost_price = cost_price or existing.cost_price
                existing.selling_price = selling_price or existing.selling_price
                if stock_qty:
                    existing.stock_qty = stock_qty
                existing.reorder_level = reorder_level
                existing.unit = unit
                existing.status = status
                if cat_id:
                    existing.category_id = cat_id
                if barcode:
                    existing.barcode = barcode
                if description:
                    existing.description = description
                updated += 1
            else:
                p = Product(
                    account_id=account_id,
                    name=name,
                    sku=sku,
                    barcode=barcode,
                    category_id=cat_id,
                    cost_price=cost_price,
                    selling_price=selling_price,
                    stock_qty=stock_qty,
                    reorder_level=reorder_level,
                    unit=unit,
                    status=status,
                    description=description,
                )
                db.session.add(p)
                db.session.flush()
                if stock_qty > 0:
                    db.session.execute(
                        insert(InventoryMovement).values(
                            product_id=p.id,
                            movement_type="opening",
                            qty_before=0,
                            qty_change=stock_qty,
                            qty_after=stock_qty,
                            notes="Smart import opening stock",
                            created_by=created_by,
                            created_at=datetime.utcnow(),
                        )
                    )
                created += 1
        except Exception as exc:
            skipped += 1
            errors.append({"row": i, "error": str(exc)})

    # Remember store engine on account settings
    if account_id:
        for key, val in (
            ("pos_engine_type", store_type),
            ("pos_engine_mode", cfg.get("pos_mode", "retail")),
        ):
            row = Setting.query.filter_by(key=key, account_id=account_id).first()
            if row:
                row.value = val
            else:
                db.session.add(Setting(key=key, value=val, account_id=account_id))

    db.session.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:50],
        "categories_created": sorted(categories_created),
        "store_type": store_type,
        "total_processed": created + updated + skipped,
    }